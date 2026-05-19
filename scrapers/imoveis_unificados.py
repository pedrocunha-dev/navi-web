import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font

def executar_unificacao_planilhas():
    pasta_planilhas = Path.home() / "AppData" / "Local" / "NAVI" / "tmp" / "imobiliario"
    pasta_planilhas.mkdir(parents=True, exist_ok=True)

    arquivos_xlsx = list(pasta_planilhas.glob("*.xlsx"))

    dfs = []

    for arquivo_path in arquivos_xlsx:
        try:
            df = pd.read_excel(arquivo_path)

            nome_arquivo = arquivo_path.stem.lower()
            if "site2" in nome_arquivo:
                fonte = "site2"
            elif "site3" in nome_arquivo:
                fonte = "site3"
            elif "site1" in nome_arquivo:
                fonte = "site1"
            else:
                fonte = "unknown"

            df['Fonte'] = fonte

            nome_arquivo = arquivo_path.name
            lote_numero = nome_arquivo.split('_lote_')[-1].split('.')[0]
            df['Lote'] = int(lote_numero)

            if 'Descrição' in df.columns:
                df = df.drop(columns=['Descrição'])
            if 'Área Construída (m²)' in df.columns and 'Área (m²)' not in df.columns:
                df = df.rename(columns={'Área Construída (m²)': 'Área (m²)'})

            dfs.append(df)
        except Exception as e:
            print(f"Error reading {arquivo_path}: {e}")

    if not dfs:
        raise FileNotFoundError("No scraping runs found. There are no files to consolidate.")

    df_geral = pd.concat(dfs, ignore_index=True)

    df_geral['Preço'] = (
        df_geral['Preço']
        .astype(str)
        .str.replace(r'[^\d,]', '', regex=True)
        .str.replace('.', '', regex=False)
        .str.replace(',', '.', regex=False)
    )

    df_geral['Área (m²)'] = (
        df_geral['Área (m²)']
        .astype(str)
        .str.replace(r'[^\d,\.]', '', regex=True)
        .str.replace(',', '.', regex=False)
    )

    df_geral['Preço'] = pd.to_numeric(df_geral['Preço'], errors='coerce')
    df_geral['Área (m²)'] = pd.to_numeric(df_geral['Área (m²)'], errors='coerce')

    df_geral = df_geral.dropna(subset=['Preço', 'Área (m²)'])
    df_geral = df_geral[(df_geral['Preço'] > 0) & (df_geral['Área (m²)'] > 0)]

    df_geral['Link'] = df_geral['Link'].astype(str).str.strip()
    df_filtrado = df_geral[
        (df_geral['Link'] != '') &
        (df_geral['Link'].str.lower() != 'link não disponível')
    ]

    df_filtrado = df_filtrado.drop_duplicates(subset=['Área (m²)', 'Preço'])

    print(df_filtrado[['ID', 'Fonte', 'Lote']].head())

    def tem_imagens_disponiveis(row):
        fonte = row['Fonte'].lower().replace(" ", "_")
        lote = row['Lote']
        id_imovel = row['ID']

        pasta_imagens = pasta_planilhas / f"imagens_{fonte}_lote_{lote}"
        pasta_imovel = pasta_imagens / f"imovel_{id_imovel}"

        return pasta_imovel.exists() and any(pasta_imovel.glob("*.jpg"))

    df_filtrado = df_filtrado[df_filtrado.apply(tem_imagens_disponiveis, axis=1)].copy()

    print(f"Records after photo filter: {len(df_filtrado)}")

    df_filtrado.insert(df_filtrado.columns.get_loc('ID'), 'Nº', range(1, len(df_filtrado) + 1))

    colunas = list(df_filtrado.columns)
    colunas.remove('Link')
    colunas.append('Link')
    df_filtrado = df_filtrado[colunas]

    arquivo_saida = pasta_planilhas / "planilha_geral_preliminar.xlsx"
    df_filtrado.to_excel(arquivo_saida, index=False)

    print(f"Preliminary spreadsheet saved to: {arquivo_saida}")

    pasta_saida = Path.home() / "Downloads"
    arquivo_resumido = pasta_saida / "imoveis_unificados.xlsx"

    cols_drop = ['ID', 'Lote']
    if 'Área Terreno (m²)' in df_filtrado.columns and df_filtrado['Área Terreno (m²)'].isna().all():
        cols_drop.append('Área Terreno (m²)')
    df_resumido = df_filtrado.drop(columns=[c for c in cols_drop if c in df_filtrado.columns])

    df_resumido.to_excel(arquivo_resumido, index=False)

    wb = load_workbook(arquivo_resumido)
    ws = wb.active

    for col_num, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, column_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        header = ws[f"{col_letter}1"].value

        for cell in column_cells[1:]:
            if header == "Endereço":
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        if header == "Endereço":
            ws.column_dimensions[col_letter].width = 40
        elif header in ["Preço", "Preço Unit. (R$/m²)", "Fonte"]:
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 13

        if header == "Preço Unit. (R$/m²)":
            for cell in column_cells[1:]:
                cell.number_format = '#,##0.00'

    wb.save(arquivo_resumido)
    print(f"Consolidated spreadsheet saved to: {arquivo_resumido}")
